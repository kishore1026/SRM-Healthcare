"""Export service – Excel/CSV generation with pandas + openpyxl via BytesIO."""
import io
from datetime import date, datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.extensions import db
from app.models.patient import Patient
from app.models.visit import PatientVisit
from app.models.medicine import Medicine
from app.models.audit_log import AuditLog


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _style_header(ws):
    """Apply styled headers to the first row of a worksheet."""
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Auto-width columns
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 4, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def _df_to_excel_bytes(df, sheet_name='Sheet1'):
    """Convert a DataFrame to styled Excel bytes."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        _style_header(writer.sheets[sheet_name])
    output.seek(0)
    return output


# ---------------------------------------------------------------------------
# Export Functions
# ---------------------------------------------------------------------------

def export_patients():
    """Export all patients to an Excel file."""
    patients = Patient.query.order_by(Patient.patient_name).all()

    data = []
    for p in patients:
        data.append({
            'Patient ID': p.patient_id,
            'Name': p.patient_name,
            'Age': p.age,
            'Gender': p.gender,
            'Designation': p.designation,
            'Student ID': p.student_id or '',
            'Staff ID': p.staff_id or '',
            'Hostel': p.hostel_name or '',
            'Room': p.room_number or '',
            'Medical History': p.medical_history or '',
            'Drug Allergy': p.drug_allergy or '',
            'Registered': p.created_at.strftime('%d-%b-%Y') if p.created_at else '',
        })

    df = pd.DataFrame(data)
    return _df_to_excel_bytes(df, 'Patients')


def export_visits(date_from=None, date_to=None):
    """Export visit history with optional date filtering."""
    query = db.session.query(PatientVisit, Patient).join(
        Patient, PatientVisit.patient_id == Patient.patient_id
    )

    if date_from:
        query = query.filter(PatientVisit.visit_date >= date_from)
    if date_to:
        query = query.filter(PatientVisit.visit_date <= date_to)

    results = query.order_by(PatientVisit.visit_date.desc()).all()

    data = []
    for visit, patient in results:
        data.append({
            'Visit ID': visit.visit_id,
            'Date': visit.visit_date.strftime('%d-%b-%Y') if visit.visit_date else '',
            'Time': visit.visit_time.strftime('%I:%M %p') if visit.visit_time else '',
            'Serial #': visit.serial_number or '',
            'Patient Name': patient.patient_name,
            'Age': patient.age,
            'Gender': patient.gender,
            'Complaint': visit.complaint or '',
            'Diagnosis': visit.diagnosis or '',
            'Treatment': visit.treatment or '',
        })

    df = pd.DataFrame(data)
    return _df_to_excel_bytes(df, 'Visit History')


def export_medicines():
    """Export full medicine inventory."""
    medicines = Medicine.query.order_by(Medicine.medicine_name).all()

    data = []
    for m in medicines:
        stock_text, _ = m.stock_status
        expiry_text, _ = m.expiry_status
        data.append({
            'Medicine ID': m.medicine_id,
            'Name': m.medicine_name,
            'Drug Type': m.drug_type,
            'Batch #': m.batch_number or '',
            'Expiry Date': m.expiry_date.strftime('%d-%b-%Y') if m.expiry_date else '',
            'Days Until Expiry': m.days_until_expiry,
            'Stock Added': m.stock_added,
            'Stock Deducted': m.stock_deducted,
            'Available Balance': m.available_balance,
            'Stock Status': stock_text,
            'Expiry Status': expiry_text,
        })

    df = pd.DataFrame(data)
    return _df_to_excel_bytes(df, 'Medicine Inventory')


def export_expiry_report():
    """Export medicines expiring soon or already expired."""
    from app.services.medicine_service import get_expiring_medicines, get_expired_medicines

    expiring = get_expiring_medicines()
    expired = get_expired_medicines()

    data = []
    for m in expired:
        data.append({
            'Name': m.medicine_name,
            'Drug Type': m.drug_type,
            'Batch #': m.batch_number or '',
            'Expiry Date': m.expiry_date.strftime('%d-%b-%Y'),
            'Days Overdue': abs(m.days_until_expiry),
            'Balance': m.available_balance,
            'Status': 'EXPIRED',
        })
    for m in expiring:
        data.append({
            'Name': m.medicine_name,
            'Drug Type': m.drug_type,
            'Batch #': m.batch_number or '',
            'Expiry Date': m.expiry_date.strftime('%d-%b-%Y'),
            'Days Remaining': m.days_until_expiry,
            'Balance': m.available_balance,
            'Status': 'EXPIRING SOON',
        })

    df = pd.DataFrame(data)
    return _df_to_excel_bytes(df, 'Expiry Report')


def export_audit_logs(date_from=None, date_to=None):
    """Export audit logs with optional date filtering."""
    query = AuditLog.query

    if date_from:
        query = query.filter(AuditLog.created_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        query = query.filter(AuditLog.created_at <= datetime.combine(date_to, datetime.max.time()))

    logs = query.order_by(AuditLog.created_at.desc()).all()

    data = []
    for log in logs:
        doctor_name = log.doctor.doctor_name if log.doctor else 'System'
        data.append({
            'Log ID': log.log_id,
            'Timestamp': log.created_at.strftime('%d-%b-%Y %I:%M %p') if log.created_at else '',
            'Doctor': doctor_name,
            'Action': log.action,
            'Module': log.module,
            'Record ID': log.record_id or '',
            'Details': log.details or '',
            'IP Address': log.ip_address or '',
        })

    df = pd.DataFrame(data)
    return _df_to_excel_bytes(df, 'Audit Logs')
